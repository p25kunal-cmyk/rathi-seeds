import openpyxl

wb = openpyxl.load_workbook('/Users/kunaljakhotiya/Downloads/Unnati & Samrudhi 2026.xlsx', data_only=False)
sheet = wb['Final Booking']

print("Formulas in AP25:AP39:")
for row in range(25, 40):
    label = sheet.cell(row=row, column=41).value  # AO
    formula = sheet.cell(row=row, column=42).value # AP
    print(f"Row {row} - {label}: {formula}")

